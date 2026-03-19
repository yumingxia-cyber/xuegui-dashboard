"""数据透视表页面"""
import streamlit as st
import pandas as pd
import numpy as np
import os
from modules import config
from modules.theme import render_header, render_kpis

st.set_page_config(page_title='数据透视表', page_icon='📋', layout='wide')
render_header('数据透视表', '多维度数据筛选与汇总分析', '📋')

OUTPUT_FILE = 'output/data_年级_日更.xlsx'
if not os.path.exists(OUTPUT_FILE):
    st.warning('尚无数据，请先到「数据上传」页面运行 pipeline')
    st.stop()

@st.cache_data
def load_data(file_mtime):
    df = pd.read_excel(OUTPUT_FILE, sheet_name='原始数据')
    df[config.DATE_COLUMN] = pd.to_datetime(df[config.DATE_COLUMN])
    return df

df = load_data(os.path.getmtime(OUTPUT_FILE))


def filter_with_toggle(label, icon, options, key):
    """带全选/清空按钮的 multiselect"""
    st.markdown(f'<p style="font-size:11px; color:#6B8DB5; margin:8px 0 2px 0;">{icon} {label}</p>', unsafe_allow_html=True)
    col_a, col_b = st.columns(2)
    with col_a:
        if st.button('全选', key=f'{key}_all', use_container_width=True):
            st.session_state[key] = list(options)
    with col_b:
        if st.button('清空', key=f'{key}_none', use_container_width=True):
            st.session_state[key] = []
    return st.multiselect(label, options, default=st.session_state.get(key, list(options)),
                           key=key, label_visibility='collapsed')


# === 侧边栏筛选器 ===
with st.sidebar:
    st.markdown('<p style="font-size:12px; font-weight:600; color:#8FB8DE; letter-spacing:0.5px; margin-bottom:8px;">FILTERS</p>', unsafe_allow_html=True)

    date_min = df[config.DATE_COLUMN].min().date()
    date_max = df[config.DATE_COLUMN].max().date()
    st.markdown('<p style="font-size:11px; color:#6B8DB5; margin:0 0 2px 0;">📅 日期范围</p>', unsafe_allow_html=True)
    date_range = st.date_input('日期范围', value=(date_min, date_max),
                                min_value=date_min, max_value=date_max, label_visibility='collapsed')

    planner_ids = sorted(df['规划师id'].unique())
    selected_ids = filter_with_toggle('规划师id', '👤', planner_ids, 'filter_planner')

    categories = sorted(df['对应品类'].unique())
    selected_cats = filter_with_toggle('对应品类', '📂', categories, 'filter_cat')

    months = sorted(df['月'].unique())
    selected_months = filter_with_toggle('月', '📆', months, 'filter_month')

    grades = sorted(df['年级'].unique())
    selected_grades = filter_with_toggle('年级', '🎓', grades, 'filter_grade')

    identities = sorted(df['身份'].unique())
    selected_identities = filter_with_toggle('身份', '🏷️', identities, 'filter_identity')

# 应用筛选（空列表表示不筛选该维度）
date_start, date_end = (date_range if len(date_range) == 2 else (date_min, date_max))
mask = (
    (df[config.DATE_COLUMN].dt.date >= date_start) &
    (df[config.DATE_COLUMN].dt.date <= date_end)
)
if selected_ids:
    mask = mask & df['规划师id'].isin(selected_ids)
if selected_cats:
    mask = mask & df['对应品类'].isin(selected_cats)
if selected_months:
    mask = mask & df['月'].isin(selected_months)
if selected_grades:
    mask = mask & df['年级'].isin(selected_grades)
if selected_identities:
    mask = mask & df['身份'].isin(selected_identities)
filtered = df[mask]

st.caption(f'筛选后: {len(filtered):,} 行 / 总计 {len(df):,} 行')

if len(filtered) == 0:
    st.info('当前筛选条件下无数据')
    st.stop()

# === 汇总指标 KPI ===
sums = {field: filtered[field].sum() for field in config.PIVOT_SUM_FIELDS}

icon_map = {'曝光': '👁️', '完成': '✅', '展示': '📄', '线索': '🎯', '好友': '🤝', '到课': '📚', '用户': '👤', '订单': '🛒', 'gmv': '💰'}
color_cycle = ['blue', 'green', 'teal', 'green', 'purple', 'indigo', 'orange', 'orange', 'blue']

kpi_items = []
for i, (k, v) in enumerate(sums.items()):
    icon = next((ic for key, ic in icon_map.items() if key in k), '📊')
    fmt_val = f'¥{v:,.0f}' if 'gmv' in k else f'{v:,.0f}'
    kpi_items.append((k, fmt_val, color_cycle[i % len(color_cycle)], icon))

st.markdown('<div class="panel"><div class="panel-title">📊 汇总指标</div>', unsafe_allow_html=True)
render_kpis(kpi_items, cols_per_row=5)
st.markdown('</div>', unsafe_allow_html=True)

# === 效率指标 KPI ===
calc_items = []
calc_colors = ['blue', 'green', 'purple', 'orange', 'teal']
calc_icons = ['📈', '🎯', '🤝', '⚡', '💎']
for i, (name, numerator, denominator, fmt) in enumerate(config.PIVOT_CALC_FIELDS):
    total_num = filtered[numerator].sum()
    total_den = filtered[denominator].sum()
    val = total_num / total_den if total_den != 0 else 0
    display = f'{val * 100:.1f}%' if fmt == 'pct' else f'{val:.1f}'
    calc_items.append((name, display, calc_colors[i], calc_icons[i]))

st.markdown('<div class="panel"><div class="panel-title">⚡ 效率指标</div>', unsafe_allow_html=True)
render_kpis(calc_items, cols_per_row=5)
st.markdown('</div>', unsafe_allow_html=True)

# === 明细数据（转置展示）===
st.markdown('<div class="panel"><div class="panel-title">📄 明细数据</div>', unsafe_allow_html=True)

group_options = ['规划师id', '对应品类', '月', config.DATE_COLUMN, '年级', '身份']
group_by = st.multiselect('分组维度', group_options, default=['规划师id'])

if group_by:
    pivot = filtered.groupby(group_by)[config.PIVOT_SUM_FIELDS].sum().reset_index()
    for name, numerator, denominator, fmt in config.PIVOT_CALC_FIELDS:
        pivot[name] = np.where(pivot[denominator] != 0, pivot[numerator] / pivot[denominator], 0)
        if fmt == 'pct':
            pivot[name] = (pivot[name] * 100).round(1)
        else:
            pivot[name] = pivot[name].round(1)

    # 转置：指标名作为行，分组值作为列
    group_key = pivot[group_by].astype(str).agg(' / '.join, axis=1)
    metric_cols = config.PIVOT_SUM_FIELDS + [f[0] for f in config.PIVOT_CALC_FIELDS]
    transposed = pivot[metric_cols].T
    transposed.columns = group_key
    transposed.index.name = '指标'
    st.dataframe(transposed, use_container_width=True)
    pivot_df = pivot
else:
    st.dataframe(filtered, use_container_width=True, hide_index=True)
    pivot_df = None

st.markdown('</div>', unsafe_allow_html=True)

# === 下载 Excel（底表 + 数据透视表）===
from io import BytesIO
from modules.pivot_table import build_pivot, build_summary

buffer = BytesIO()
with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
    # Sheet1: 筛选后的底表
    filtered.to_excel(writer, sheet_name='底表数据', index=False)

    # Sheet2: 数据透视表（基于当前筛选数据）
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = writer.book
    ws = wb.create_sheet(title='数据透视表')

    # 汇总区域
    summary = build_summary(filtered)
    ws['A1'] = '汇总指标'
    ws['A1'].font = Font(bold=True, size=12)
    row_idx = 2
    for key, val in summary.items():
        ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        cell = ws.cell(row=row_idx, column=2, value=val)
        for name, _, _, fmt in config.PIVOT_CALC_FIELDS:
            if key == name and fmt == 'pct':
                cell.number_format = '0.0"%"'
            elif key == name and fmt == 'num':
                cell.number_format = '0.0'
        row_idx += 1

    # 明细透视表
    full_pivot = build_pivot(filtered)
    start_row = row_idx + 1
    ws.cell(row=start_row, column=1, value='明细数据').font = Font(bold=True, size=12)
    start_row += 1

    headers = list(full_pivot.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4472C4')
        cell.alignment = Alignment(horizontal='center')

    for r_idx, row_data in enumerate(full_pivot.values, start_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            col_name = headers[c_idx - 1]
            for name, _, _, fmt in config.PIVOT_CALC_FIELDS:
                if col_name == name:
                    cell.number_format = '0.0"%"' if fmt == 'pct' else '0.0'

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = 18

buffer.seek(0)
st.download_button('📥 下载 Excel（底表 + 数据透视表）', data=buffer,
                   file_name='学规数据分析.xlsx',
                   mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
