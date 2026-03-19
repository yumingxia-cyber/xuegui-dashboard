import pandas as pd
import numpy as np
from . import config


def build_pivot(df):
    """根据筛选项生成数据透视表 DataFrame"""
    pivot = df.groupby(config.PIVOT_FILTERS)[config.PIVOT_SUM_FIELDS].sum().reset_index()

    for name, numerator, denominator, fmt in config.PIVOT_CALC_FIELDS:
        pivot[name] = np.where(
            pivot[denominator] != 0,
            pivot[numerator] / pivot[denominator],
            0
        )
        if fmt == 'pct':
            pivot[name] = (pivot[name] * 100).round(1)
        else:
            pivot[name] = pivot[name].round(1)

    return pivot


def build_summary(df):
    """生成总计汇总行（不分组，直接求和），与参考表 Sheet1 结构一致"""
    summary = {}
    for field in config.PIVOT_SUM_FIELDS:
        summary[f'求和项:{field}'] = df[field].sum()

    for name, numerator, denominator, fmt in config.PIVOT_CALC_FIELDS:
        total_num = df[numerator].sum()
        total_den = df[denominator].sum()
        val = total_num / total_den if total_den != 0 else 0
        if fmt == 'pct':
            summary[name] = round(val * 100, 1)
        else:
            summary[name] = round(val, 1)

    return summary


def write_pivot_sheet(writer, df):
    """将数据透视表写入 Excel 的 sheet"""
    from openpyxl.styles import Font, PatternFill, Alignment, numbers

    pivot = build_pivot(df)
    summary = build_summary(df)

    workbook = writer.book
    ws = workbook.create_sheet(title='数据透视表')

    # === 筛选项区域（参照参考表 Sheet1 布局）===
    header_font = Font(bold=True)
    ws['A1'] = '规划师id'
    ws['B1'] = '(全部)'
    ws['A2'] = '对应品类'
    ws['B2'] = '(全部)'
    ws['A3'] = '月'
    ws['B3'] = '(全部)'
    for row in range(1, 4):
        ws.cell(row=row, column=1).font = header_font

    # === 汇总区域 ===
    ws['A5'] = '值'
    ws['A5'].font = header_font
    row_idx = 6
    for key, val in summary.items():
        ws.cell(row=row_idx, column=1, value=key)
        cell = ws.cell(row=row_idx, column=2, value=val)
        # 百分比字段加 % 后缀显示
        for name, _, _, fmt in config.PIVOT_CALC_FIELDS:
            if key == name and fmt == 'pct':
                cell.number_format = '0.0"%"'
            elif key == name and fmt == 'num':
                cell.number_format = '0.0'
        row_idx += 1

    # === 明细数据区域 ===
    start_row = row_idx + 2
    ws.cell(row=start_row - 1, column=1, value='明细数据').font = Font(bold=True, size=12)

    # 写表头
    headers = list(pivot.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='4472C4')
        cell.alignment = Alignment(horizontal='center')

    # 写数据
    for r_idx, row_data in enumerate(pivot.values, start_row + 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            # 计算字段的格式
            col_name = headers[c_idx - 1]
            for name, _, _, fmt in config.PIVOT_CALC_FIELDS:
                if col_name == name:
                    if fmt == 'pct':
                        cell.number_format = '0.0"%"'
                    else:
                        cell.number_format = '0.0'

    # 列宽自适应
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18


def run(df, writer):
    """执行透视表模块"""
    write_pivot_sheet(writer, df)
    print("pivot_table 完成: 数据透视表已写入")
